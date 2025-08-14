#!/usr/bin/env python3
"""
African Compliance Automation Engine
Advanced automated compliance assessment for African regulatory frameworks

Features:
- POPIA (Protection of Personal Information Act) - South Africa
- NDPR (Nigeria Data Protection Regulation) - Nigeria
- DPA (Data Protection Act) - Kenya, Ghana
- CBN Guidelines - Nigeria Central Bank
- SARB Regulations - South African Reserve Bank
- EAC Data Protection Guidelines - East African Community
- AU Convention on Cyber Security - African Union
- Local banking and financial regulations
- Mobile money compliance frameworks
- Telecommunications regulations
- Mining and natural resources compliance
"""

import asyncio
import json
import logging
import xml.etree.ElementTree as ET
from typing import Dict, List, Optional, Any, Union, Tuple
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta
from pathlib import Path
import re
import hashlib
import base64
from enum import Enum
import pandas as pd
import numpy as np

# Compliance and audit libraries
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("Excel libraries not available - report generation limited")

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("PDF libraries not available - report generation limited")

from .base_service import SecurityService, ServiceHealth, ServiceStatus

logger = logging.getLogger(__name__)

class ComplianceStatus(Enum):
    COMPLIANT = "compliant"
    NON_COMPLIANT = "non_compliant"
    PARTIALLY_COMPLIANT = "partially_compliant"
    NOT_APPLICABLE = "not_applicable"
    UNDER_REVIEW = "under_review"

class RiskLevel(Enum):
    CRITICAL = "critical"
    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    NEGLIGIBLE = "negligible"

class ComplianceFramework(Enum):
    POPIA = "popia"
    NDPR = "ndpr"
    KENYA_DPA = "kenya_dpa"
    GHANA_DPA = "ghana_dpa"
    CBN_GUIDELINES = "cbn_guidelines"
    SARB_REGULATIONS = "sarb_regulations"
    EAC_GUIDELINES = "eac_guidelines"
    AU_CONVENTION = "au_convention"
    PCI_DSS = "pci_dss"
    ISO_27001 = "iso_27001"

@dataclass
class ComplianceRequirement:
    """Individual compliance requirement"""
    requirement_id: str
    framework: ComplianceFramework
    category: str
    title: str
    description: str
    mandatory: bool
    control_objective: str
    implementation_guidance: str
    assessment_criteria: List[str]
    evidence_requirements: List[str]
    penalty_information: Dict[str, Any]
    related_requirements: List[str]
    last_updated: datetime

@dataclass
class ComplianceAssessmentResult:
    """Individual compliance assessment result"""
    assessment_id: str
    requirement_id: str
    status: ComplianceStatus
    compliance_score: float
    risk_level: RiskLevel
    findings: List[str]
    evidence_collected: List[str]
    gaps_identified: List[str]
    remediation_actions: List[str]
    estimated_cost: float
    timeline_to_compliance: int  # days
    business_impact: str
    technical_implementation: str
    assessment_date: datetime
    assessor_notes: str

@dataclass
class ComplianceReport:
    """Comprehensive compliance report"""
    report_id: str
    organization_name: str
    assessment_scope: str
    frameworks_assessed: List[ComplianceFramework]
    overall_compliance_score: float
    overall_risk_level: RiskLevel
    assessment_results: List[ComplianceAssessmentResult]
    executive_summary: Dict[str, Any]
    detailed_findings: Dict[str, Any]
    remediation_roadmap: List[Dict[str, Any]]
    cost_analysis: Dict[str, Any]
    regulatory_implications: Dict[str, Any]
    recommendations: List[str]
    next_assessment_date: datetime
    report_generation_date: datetime
    compliance_officer: str

class AfricanComplianceAutomationEngine(SecurityService):
    """
    Advanced Compliance Automation Engine for African Regulatory Frameworks

    Capabilities:
    - Automated compliance assessment across multiple African frameworks
    - Risk-based compliance prioritization
    - Intelligent gap analysis and remediation planning
    - Regulatory change monitoring and impact assessment
    - Automated evidence collection and documentation
    - Multi-format reporting (PDF, Excel, JSON)
    - Integration with African regulatory databases
    - Real-time compliance monitoring and alerting
    """

    def __init__(self, **kwargs):
        super().__init__(
            service_id="african_compliance_automation",
            dependencies=["database", "redis", "vault", "document_processor"],
            config=kwargs.get("config", {})
        )

        # Compliance frameworks registry
        self.compliance_frameworks = {}
        self.requirements_database = {}
        self.assessment_templates = {}

        # Assessment state and history
        self.active_assessments = {}
        self.assessment_history = []
        self.compliance_baselines = {}

        # Regulatory intelligence
        self.regulatory_updates = {}
        self.penalty_database = {}
        self.precedent_cases = {}

        # Evidence collection and management
        self.evidence_collectors = {}
        self.documentation_templates = {}
        self.audit_trails = {}

        # African regulatory frameworks configuration
        self.african_frameworks = {
            ComplianceFramework.POPIA: {
                "name": "Protection of Personal Information Act",
                "country": "South Africa",
                "authority": "Information Regulator",
                "effective_date": datetime(2021, 7, 1),
                "scope": ["data_protection", "privacy", "information_security"],
                "penalties": {"max_fine": 10000000, "criminal_liability": True},
                "update_frequency": "quarterly"
            },
            ComplianceFramework.NDPR: {
                "name": "Nigeria Data Protection Regulation",
                "country": "Nigeria",
                "authority": "NITDA",
                "effective_date": datetime(2019, 1, 25),
                "scope": ["data_protection", "privacy", "cross_border_transfers"],
                "penalties": {"max_fine": 100000000, "criminal_liability": False},
                "update_frequency": "bi_annual"
            },
            ComplianceFramework.KENYA_DPA: {
                "name": "Data Protection Act",
                "country": "Kenya",
                "authority": "Data Protection Commissioner",
                "effective_date": datetime(2019, 11, 25),
                "scope": ["data_protection", "privacy", "digital_rights"],
                "penalties": {"max_fine": 5000000, "criminal_liability": True},
                "update_frequency": "annual"
            },
            ComplianceFramework.CBN_GUIDELINES: {
                "name": "Central Bank of Nigeria Guidelines",
                "country": "Nigeria",
                "authority": "Central Bank of Nigeria",
                "effective_date": datetime(2020, 1, 1),
                "scope": ["banking", "fintech", "mobile_money", "cybersecurity"],
                "penalties": {"max_fine": 500000000, "license_revocation": True},
                "update_frequency": "quarterly"
            }
        }

        # Industry-specific compliance mappings
        self.industry_mappings = {
            "banking": [ComplianceFramework.CBN_GUIDELINES, ComplianceFramework.SARB_REGULATIONS, ComplianceFramework.PCI_DSS],
            "fintech": [ComplianceFramework.CBN_GUIDELINES, ComplianceFramework.POPIA, ComplianceFramework.NDPR],
            "telecommunications": [ComplianceFramework.EAC_GUIDELINES, ComplianceFramework.POPIA, ComplianceFramework.NDPR],
            "healthcare": [ComplianceFramework.POPIA, ComplianceFramework.NDPR, ComplianceFramework.KENYA_DPA],
            "government": [ComplianceFramework.AU_CONVENTION, ComplianceFramework.POPIA, ComplianceFramework.EAC_GUIDELINES],
            "mining": [ComplianceFramework.POPIA, ComplianceFramework.ISO_27001, ComplianceFramework.AU_CONVENTION],
            "general": [ComplianceFramework.POPIA, ComplianceFramework.NDPR, ComplianceFramework.ISO_27001]
        }

    async def initialize(self) -> bool:
        """Initialize the African compliance automation engine"""
        try:
            logger.info("Initializing African Compliance Automation Engine...")

            # Load compliance requirements database
            await self._load_compliance_requirements()

            # Initialize assessment templates
            await self._initialize_assessment_templates()

            # Set up evidence collection systems
            await self._setup_evidence_collection()

            # Initialize regulatory monitoring
            await self._initialize_regulatory_monitoring()

            # Load penalty and precedent databases
            await self._load_penalty_database()

            # Set up automated reporting
            await self._setup_automated_reporting()

            logger.info("African Compliance Automation Engine initialized successfully")
            return True

        except Exception as e:
            logger.error(f"Failed to initialize compliance automation engine: {e}")
            return False

    async def conduct_comprehensive_compliance_assessment(self,
                                                        organization_data: Dict[str, Any],
                                                        assessment_scope: str = "full") -> ComplianceReport:
        """Conduct comprehensive compliance assessment"""
        try:
            report_id = f"compliance_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            start_time = datetime.now()

            # Determine applicable frameworks
            applicable_frameworks = await self._determine_applicable_frameworks(organization_data)

            # Initialize assessment results
            assessment_results = []

            # Assess each applicable framework
            for framework in applicable_frameworks:
                framework_results = await self._assess_framework_compliance(
                    framework, organization_data, assessment_scope
                )
                assessment_results.extend(framework_results)

            # Calculate overall compliance metrics
            overall_metrics = await self._calculate_overall_compliance_metrics(assessment_results)

            # Generate executive summary
            executive_summary = await self._generate_executive_summary(
                organization_data, assessment_results, overall_metrics
            )

            # Identify detailed findings
            detailed_findings = await self._analyze_detailed_findings(assessment_results)

            # Create remediation roadmap
            remediation_roadmap = await self._create_remediation_roadmap(
                assessment_results, organization_data
            )

            # Perform cost analysis
            cost_analysis = await self._perform_cost_analysis(
                remediation_roadmap, organization_data
            )

            # Analyze regulatory implications
            regulatory_implications = await self._analyze_regulatory_implications(
                assessment_results, applicable_frameworks
            )

            # Generate recommendations
            recommendations = await self._generate_strategic_recommendations(
                assessment_results, cost_analysis, regulatory_implications
            )

            # Create comprehensive report
            compliance_report = ComplianceReport(
                report_id=report_id,
                organization_name=organization_data.get("organization_name", "Unknown"),
                assessment_scope=assessment_scope,
                frameworks_assessed=applicable_frameworks,
                overall_compliance_score=overall_metrics["compliance_score"],
                overall_risk_level=overall_metrics["risk_level"],
                assessment_results=assessment_results,
                executive_summary=executive_summary,
                detailed_findings=detailed_findings,
                remediation_roadmap=remediation_roadmap,
                cost_analysis=cost_analysis,
                regulatory_implications=regulatory_implications,
                recommendations=recommendations,
                next_assessment_date=await self._calculate_next_assessment_date(applicable_frameworks),
                report_generation_date=datetime.now(),
                compliance_officer=organization_data.get("compliance_officer", "System Generated")
            )

            # Store assessment for tracking
            self.assessment_history.append(compliance_report)

            logger.info(f"Comprehensive compliance assessment completed: {overall_metrics['compliance_score']:.2f}% compliant")
            return compliance_report

        except Exception as e:
            logger.error(f"Comprehensive compliance assessment failed: {e}")
            raise

    async def assess_popia_compliance(self, organization_data: Dict[str, Any]) -> List[ComplianceAssessmentResult]:
        """Detailed POPIA (Protection of Personal Information Act) compliance assessment"""
        try:
            popia_results = []

            # POPIA Chapter 3: Processing of Personal Information
            chapter3_results = await self._assess_popia_chapter3(organization_data)
            popia_results.extend(chapter3_results)

            # POPIA Chapter 4: Responsibilities of Responsible Parties
            chapter4_results = await self._assess_popia_chapter4(organization_data)
            popia_results.extend(chapter4_results)

            # POPIA Chapter 5: Operator Requirements
            chapter5_results = await self._assess_popia_chapter5(organization_data)
            popia_results.extend(chapter5_results)

            # POPIA Chapter 6: Direct Marketing
            chapter6_results = await self._assess_popia_chapter6(organization_data)
            popia_results.extend(chapter6_results)

            # POPIA Chapter 7: Transborder Information Flows
            chapter7_results = await self._assess_popia_chapter7(organization_data)
            popia_results.extend(chapter7_results)

            logger.info(f"POPIA compliance assessment completed: {len(popia_results)} requirements assessed")
            return popia_results

        except Exception as e:
            logger.error(f"POPIA compliance assessment failed: {e}")
            return []

    async def assess_ndpr_compliance(self, organization_data: Dict[str, Any]) -> List[ComplianceAssessmentResult]:
        """Detailed NDPR (Nigeria Data Protection Regulation) compliance assessment"""
        try:
            ndpr_results = []

            # NDPR Part II: Lawfulness of Processing
            part2_results = await self._assess_ndpr_lawfulness(organization_data)
            ndpr_results.extend(part2_results)

            # NDPR Part III: Data Subject Rights
            part3_results = await self._assess_ndpr_data_subject_rights(organization_data)
            ndpr_results.extend(part3_results)

            # NDPR Part IV: Data Controller and Processor Obligations
            part4_results = await self._assess_ndpr_controller_obligations(organization_data)
            ndpr_results.extend(part4_results)

            # NDPR Part V: Transfer of Personal Data
            part5_results = await self._assess_ndpr_data_transfers(organization_data)
            ndpr_results.extend(part5_results)

            # NDPR Part VI: Data Protection Officer
            part6_results = await self._assess_ndpr_dpo_requirements(organization_data)
            ndpr_results.extend(part6_results)

            # NDPR Part VII: Data Protection Audit
            part7_results = await self._assess_ndpr_audit_requirements(organization_data)
            ndpr_results.extend(part7_results)

            logger.info(f"NDPR compliance assessment completed: {len(ndpr_results)} requirements assessed")
            return ndpr_results

        except Exception as e:
            logger.error(f"NDPR compliance assessment failed: {e}")
            return []

    async def assess_cbn_guidelines_compliance(self, organization_data: Dict[str, Any]) -> List[ComplianceAssessmentResult]:
        """Detailed CBN Guidelines compliance assessment"""
        try:
            cbn_results = []

            # Cybersecurity Framework
            cybersecurity_results = await self._assess_cbn_cybersecurity_framework(organization_data)
            cbn_results.extend(cybersecurity_results)

            # Risk Management Framework
            risk_mgmt_results = await self._assess_cbn_risk_management(organization_data)
            cbn_results.extend(risk_mgmt_results)

            # IT Governance Framework
            it_governance_results = await self._assess_cbn_it_governance(organization_data)
            cbn_results.extend(it_governance_results)

            # Payment System Management
            payment_system_results = await self._assess_cbn_payment_systems(organization_data)
            cbn_results.extend(payment_system_results)

            # Consumer Protection Framework
            consumer_protection_results = await self._assess_cbn_consumer_protection(organization_data)
            cbn_results.extend(consumer_protection_results)

            # Mobile Money Guidelines
            mobile_money_results = await self._assess_cbn_mobile_money(organization_data)
            cbn_results.extend(mobile_money_results)

            logger.info(f"CBN Guidelines compliance assessment completed: {len(cbn_results)} requirements assessed")
            return cbn_results

        except Exception as e:
            logger.error(f"CBN Guidelines compliance assessment failed: {e}")
            return []

    async def automated_evidence_collection(self, requirements: List[str],
                                          organization_data: Dict[str, Any]) -> Dict[str, Any]:
        """Automated evidence collection for compliance requirements"""
        try:
            evidence_collection_result = {
                "collection_id": f"evidence_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                "requirements_processed": len(requirements),
                "evidence_collected": {},
                "automated_tests": {},
                "manual_review_required": [],
                "collection_summary": {},
                "confidence_scores": {}
            }

            for requirement_id in requirements:
                # Get requirement details
                requirement = await self._get_requirement_details(requirement_id)
                if not requirement:
                    continue

                # Automated evidence collection
                automated_evidence = await self._collect_automated_evidence(requirement, organization_data)
                evidence_collection_result["evidence_collected"][requirement_id] = automated_evidence

                # Automated compliance tests
                automated_tests = await self._run_automated_compliance_tests(requirement, organization_data)
                evidence_collection_result["automated_tests"][requirement_id] = automated_tests

                # Determine if manual review is required
                if automated_tests.get("confidence_score", 0) < 0.8:
                    evidence_collection_result["manual_review_required"].append({
                        "requirement_id": requirement_id,
                        "reason": "Low confidence in automated assessment",
                        "confidence_score": automated_tests.get("confidence_score", 0)
                    })

                # Calculate confidence score
                confidence_score = await self._calculate_evidence_confidence_score(
                    automated_evidence, automated_tests
                )
                evidence_collection_result["confidence_scores"][requirement_id] = confidence_score

            # Generate collection summary
            evidence_collection_result["collection_summary"] = await self._generate_evidence_collection_summary(
                evidence_collection_result
            )

            logger.info(f"Automated evidence collection completed: {len(requirements)} requirements processed")
            return evidence_collection_result

        except Exception as e:
            logger.error(f"Automated evidence collection failed: {e}")
            return {"error": str(e)}

    async def generate_compliance_report(self, compliance_report: ComplianceReport,
                                       output_format: str = "pdf") -> str:
        """Generate compliance report in specified format"""
        try:
            if output_format.lower() == "pdf" and PDF_AVAILABLE:
                report_path = await self._generate_pdf_report(compliance_report)
            elif output_format.lower() == "excel" and EXCEL_AVAILABLE:
                report_path = await self._generate_excel_report(compliance_report)
            elif output_format.lower() == "json":
                report_path = await self._generate_json_report(compliance_report)
            else:
                # Fallback to JSON
                report_path = await self._generate_json_report(compliance_report)

            logger.info(f"Compliance report generated: {report_path}")
            return report_path

        except Exception as e:
            logger.error(f"Compliance report generation failed: {e}")
            raise

    async def real_time_compliance_monitoring(self, organization_data: Dict[str, Any]) -> Dict[str, Any]:
        """Real-time compliance monitoring and alerting"""
        try:
            monitoring_result = {
                "monitoring_id": f"monitor_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                "compliance_status": {},
                "alerts_generated": [],
                "trend_analysis": {},
                "recommendations": [],
                "next_review_dates": {}
            }

            # Monitor each applicable framework
            applicable_frameworks = await self._determine_applicable_frameworks(organization_data)

            for framework in applicable_frameworks:
                # Real-time compliance check
                current_status = await self._check_real_time_compliance(framework, organization_data)
                monitoring_result["compliance_status"][framework.value] = current_status

                # Generate alerts for non-compliance
                if current_status.get("status") == ComplianceStatus.NON_COMPLIANT:
                    alert = await self._generate_compliance_alert(framework, current_status, organization_data)
                    monitoring_result["alerts_generated"].append(alert)

                # Calculate next review date
                next_review = await self._calculate_next_review_date(framework, current_status)
                monitoring_result["next_review_dates"][framework.value] = next_review

            # Trend analysis
            trend_analysis = await self._analyze_compliance_trends(
                monitoring_result["compliance_status"], organization_data
            )
            monitoring_result["trend_analysis"] = trend_analysis

            # Generate recommendations
            recommendations = await self._generate_monitoring_recommendations(
                monitoring_result["compliance_status"], trend_analysis
            )
            monitoring_result["recommendations"] = recommendations

            logger.info(f"Real-time compliance monitoring completed: {len(monitoring_result['alerts_generated'])} alerts generated")
            return monitoring_result

        except Exception as e:
            logger.error(f"Real-time compliance monitoring failed: {e}")
            return {"error": str(e)}

    # Implementation helper methods for specific frameworks
    async def _assess_popia_chapter3(self, organization_data: Dict[str, Any]) -> List[ComplianceAssessmentResult]:
        """Assess POPIA Chapter 3: Processing of Personal Information"""
        results = []

        # Condition 1: Accountability
        accountability_result = await self._assess_requirement(
            "popia_3_1_accountability",
            organization_data,
            {
                "title": "Accountability",
                "description": "Responsible party must ensure processing complies with POPIA",
                "assessment_criteria": [
                    "Data protection policy exists",
                    "Compliance monitoring mechanisms in place",
                    "Regular compliance reviews conducted",
                    "Accountability documentation maintained"
                ]
            }
        )
        results.append(accountability_result)

        # Condition 2: Processing Limitation
        processing_limitation_result = await self._assess_requirement(
            "popia_3_2_processing_limitation",
            organization_data,
            {
                "title": "Processing Limitation",
                "description": "Personal information must be processed lawfully and reasonably",
                "assessment_criteria": [
                    "Lawful basis identified for processing",
                    "Processing is necessary for legitimate purpose",
                    "Minimization principle applied",
                    "Retention periods defined"
                ]
            }
        )
        results.append(processing_limitation_result)

        # Continue with other conditions...

        return results

    async def _assess_ndpr_lawfulness(self, organization_data: Dict[str, Any]) -> List[ComplianceAssessmentResult]:
        """Assess NDPR Part II: Lawfulness of Processing"""
        results = []

        # Article 2.1: Lawful Basis
        lawful_basis_result = await self._assess_requirement(
            "ndpr_2_1_lawful_basis",
            organization_data,
            {
                "title": "Lawful Basis for Processing",
                "description": "Processing must be based on legitimate grounds",
                "assessment_criteria": [
                    "Consent obtained where required",
                    "Contract processing justified",
                    "Legal obligation documented",
                    "Vital interests identified"
                ]
            }
        )
        results.append(lawful_basis_result)

        # Continue with other articles...

        return results

    async def _assess_cbn_cybersecurity_framework(self, organization_data: Dict[str, Any]) -> List[ComplianceAssessmentResult]:
        """Assess CBN Cybersecurity Framework requirements"""
        results = []

        # Governance and Risk Management
        governance_result = await self._assess_requirement(
            "cbn_cyber_governance",
            organization_data,
            {
                "title": "Cybersecurity Governance",
                "description": "Establish cybersecurity governance framework",
                "assessment_criteria": [
                    "Board oversight of cybersecurity",
                    "CISO appointment and reporting",
                    "Cybersecurity policies approved",
                    "Risk assessment conducted"
                ]
            }
        )
        results.append(governance_result)

        # Continue with other framework elements...

        return results

    # Report generation methods
    async def _generate_pdf_report(self, compliance_report: ComplianceReport) -> str:
        """Generate PDF compliance report"""
        try:
            if not PDF_AVAILABLE:
                raise Exception("PDF generation libraries not available")

            report_filename = f"compliance_report_{compliance_report.report_id}.pdf"
            report_path = Path(f"/tmp/{report_filename}")

            doc = SimpleDocTemplate(str(report_path), pagesize=A4)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title = Paragraph(f"Compliance Assessment Report", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))

            # Executive Summary
            exec_summary = Paragraph("Executive Summary", styles['Heading1'])
            story.append(exec_summary)

            summary_text = f"""
            Organization: {compliance_report.organization_name}
            Assessment Date: {compliance_report.report_generation_date.strftime('%Y-%m-%d')}
            Overall Compliance Score: {compliance_report.overall_compliance_score:.1f}%
            Risk Level: {compliance_report.overall_risk_level.value.title()}
            Frameworks Assessed: {', '.join([f.value for f in compliance_report.frameworks_assessed])}
            """

            story.append(Paragraph(summary_text, styles['Normal']))
            story.append(Spacer(1, 12))

            # Detailed findings table
            findings_data = [['Requirement', 'Status', 'Risk Level', 'Score']]
            for result in compliance_report.assessment_results[:20]:  # Limit for PDF
                findings_data.append([
                    result.requirement_id,
                    result.status.value,
                    result.risk_level.value,
                    f"{result.compliance_score:.1f}%"
                ])

            findings_table = Table(findings_data)
            findings_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            story.append(findings_table)

            # Build PDF
            doc.build(story)

            return str(report_path)

        except Exception as e:
            logger.error(f"PDF report generation failed: {e}")
            raise

    async def _generate_excel_report(self, compliance_report: ComplianceReport) -> str:
        """Generate Excel compliance report"""
        try:
            if not EXCEL_AVAILABLE:
                raise Exception("Excel generation libraries not available")

            report_filename = f"compliance_report_{compliance_report.report_id}.xlsx"
            report_path = Path(f"/tmp/{report_filename}")

            workbook = openpyxl.Workbook()

            # Executive Summary Sheet
            exec_sheet = workbook.active
            exec_sheet.title = "Executive Summary"

            # Headers
            exec_sheet['A1'] = "African Compliance Assessment Report"
            exec_sheet['A1'].font = Font(size=16, bold=True)

            # Summary data
            exec_sheet['A3'] = "Organization:"
            exec_sheet['B3'] = compliance_report.organization_name
            exec_sheet['A4'] = "Assessment Date:"
            exec_sheet['B4'] = compliance_report.report_generation_date.strftime('%Y-%m-%d')
            exec_sheet['A5'] = "Overall Compliance Score:"
            exec_sheet['B5'] = f"{compliance_report.overall_compliance_score:.1f}%"
            exec_sheet['A6'] = "Risk Level:"
            exec_sheet['B6'] = compliance_report.overall_risk_level.value.title()

            # Detailed Results Sheet
            results_sheet = workbook.create_sheet(title="Detailed Results")

            headers = ['Requirement ID', 'Status', 'Score', 'Risk Level', 'Findings', 'Remediation Actions']
            for col, header in enumerate(headers, 1):
                cell = results_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Add results data
            for row, result in enumerate(compliance_report.assessment_results, 2):
                results_sheet.cell(row=row, column=1, value=result.requirement_id)
                results_sheet.cell(row=row, column=2, value=result.status.value)
                results_sheet.cell(row=row, column=3, value=f"{result.compliance_score:.1f}%")
                results_sheet.cell(row=row, column=4, value=result.risk_level.value)
                results_sheet.cell(row=row, column=5, value='; '.join(result.findings))
                results_sheet.cell(row=row, column=6, value='; '.join(result.remediation_actions))

            # Save workbook
            workbook.save(report_path)

            return str(report_path)

        except Exception as e:
            logger.error(f"Excel report generation failed: {e}")
            raise

    # Additional helper methods would be implemented here...

    async def health_check(self) -> ServiceHealth:
        """Health check for African compliance automation engine"""
        try:
            checks = {
                "frameworks_loaded": len(self.compliance_frameworks),
                "requirements_database_size": len(self.requirements_database),
                "assessment_templates_ready": len(self.assessment_templates),
                "pdf_generation_available": PDF_AVAILABLE,
                "excel_generation_available": EXCEL_AVAILABLE,
                "active_assessments": len(self.active_assessments),
                "regulatory_updates_current": len(self.regulatory_updates)
            }

            status = ServiceStatus.HEALTHY if all([
                checks["frameworks_loaded"] > 0,
                checks["requirements_database_size"] > 0
            ]) else ServiceStatus.DEGRADED

            return ServiceHealth(
                service_id=self.service_id,
                status=status,
                checks=checks,
                timestamp=datetime.utcnow()
            )

        except Exception as e:
            return ServiceHealth(
                service_id=self.service_id,
                status=ServiceStatus.UNHEALTHY,
                checks={"error": str(e)},
                timestamp=datetime.utcnow()
            )

# Export the African compliance automation engine
__all__ = [
    "AfricanComplianceAutomationEngine",
    "ComplianceRequirement",
    "ComplianceAssessmentResult",
    "ComplianceReport",
    "ComplianceStatus",
    "RiskLevel",
    "ComplianceFramework"
]
